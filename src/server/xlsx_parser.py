import openpyxl


def parse_xlsx_character(file_path: str) -> dict:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    for ws_name in ["人物卡", "简化卡", "Sheet1"]:
        if ws_name in wb.sheetnames:
            ws = wb[ws_name]
            return _parse_sheet(ws)

    ws = wb.active
    return _parse_sheet(ws)


def _parse_sheet(ws) -> dict:
    data = {}
    for row in ws.iter_rows(min_row=1, max_row=50, values_only=False):
        if len(row) >= 2 and row[0].value is not None and row[1].value is not None:
            key = str(row[0].value).strip()
            val = row[1].value
            data[key] = val

    name = data.get("姓名", data.get("Name", "未知"))
    hp = _to_int(data.get("HP", data.get("生命", 10)))
    san = _to_int(data.get("SAN", data.get("理智", 50)))
    mp = _to_int(data.get("MP", data.get("魔法", 10)))
    luck = _to_int(data.get("LUCK", data.get("幸运", 50)))

    skills = {}
    skip_keys = {"HP", "SAN", "MP", "LUCK", "生命", "理智", "魔法", "幸运", "姓名", "Name"}
    for k, v in data.items():
        if k not in skip_keys and isinstance(v, (int, float)):
            skills[k] = int(v)

    return {
        "name": str(name),
        "hp": hp,
        "san": san,
        "mp": mp,
        "luck": luck,
        "skills": skills,
        "raw": data,
    }


def _to_int(val) -> int:
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return 0
