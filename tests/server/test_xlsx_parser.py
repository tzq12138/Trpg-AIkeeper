import openpyxl


def _make_xlsx(path: str, data: dict[str, str | int]):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "人物卡"
    for i, (k, v) in enumerate(data.items(), start=1):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
    wb.save(path)


def test_parse_xlsx_returns_character_fields(tmp_path):
    from src.server.xlsx_parser import parse_xlsx_character

    path = str(tmp_path / "test.xlsx")
    _make_xlsx(path, {"姓名": "张三", "HP": 12, "SAN": 65, "MP": 10, "LUCK": 50})

    result = parse_xlsx_character(path)
    assert result["name"] == "张三"
    assert result["hp"] == 12
    assert result["san"] == 65
    assert result["mp"] == 10
    assert result["luck"] == 50


def test_parse_xlsx_with_skills(tmp_path):
    from src.server.xlsx_parser import parse_xlsx_character

    path = str(tmp_path / "test.xlsx")
    _make_xlsx(path, {"姓名": "李四", "HP": 10, "SAN": 50, "图书馆使用": 40, "聆听": 35})

    result = parse_xlsx_character(path)
    assert result["name"] == "李四"
    assert result["skills"]["图书馆使用"] == 40
    assert result["skills"]["聆听"] == 35


def test_parse_xlsx_missing_fields(tmp_path):
    from src.server.xlsx_parser import parse_xlsx_character

    path = str(tmp_path / "test.xlsx")
    _make_xlsx(path, {"姓名": "王五"})

    result = parse_xlsx_character(path)
    assert result["name"] == "王五"
    assert result["hp"] == 10  # default
    assert result["san"] == 50  # default
